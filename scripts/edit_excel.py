from datetime import datetime
from openpyxl import load_workbook

PATH = "gpt.xlsx"
SHEET = "Tabellenblatt1"
CELL = "A1"

wb = load_workbook(PATH)
ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

ws[CELL] = f"Updated by GitHub Actions: {datetime.utcnow().isoformat()}Z"
wb.save(PATH)
print("Saved changes to", PATH)
